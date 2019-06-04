from openpyxl import Workbook, load_workbook
import os


all_files = list()
rootdir = os.getcwd()

for root, dirs, files in os.walk(rootdir):
    for file in files:
        if file.endswith(".xlsx"):
             all_files.append(os.path.join(root, file))

arr = ['nomer', 'okrug', 'nazvanie_ac','nazvanie_filiala','adres_sdaniya',
'familiya','imya','otchestvo',
'fio','summa_stavok','stoyka_info','medpost','priem s vrachom','manipulyation','kabinet_llo',
'procedurnaya','storage']


wb_new=Workbook()

filepath="{}/final.xlsx".format(rootdir)

wb_new.save(filepath)



wb_final=load_workbook(filepath)

sheet2=wb_final.active


for i in range(len(arr)):
	sheet2.cell(row=1, column=i+1).value = arr[i]


line = 1
for i in all_files:
	wb = load_workbook(i)
	print(wb.get_sheet_names())
	sheet = wb['Время']
	max_row=sheet.max_row
	for row in range(max_row-2):
		line = line+1
		sheet2.cell(row=line, column=1).value = line - 1
		sheet2.cell(row=line, column=2).value = sheet.cell(row=row+2, column=2).value
		sheet2.cell(row=line, column=3).value = sheet.cell(row=row+2, column=3).value
		sheet2.cell(row=line, column=4).value = sheet.cell(row=row+2, column=4).value
		sheet2.cell(row=line, column=5).value = sheet.cell(row=row+2, column=5).value
		sheet2.cell(row=line, column=6).value = sheet.cell(row=row+2, column=6).value
		sheet2.cell(row=line, column=7).value = sheet.cell(row=row+2, column=7).value
		sheet2.cell(row=line, column=8).value = sheet.cell(row=row+2, column=8).value
		sheet2.cell(row=line, column=9).value = sheet.cell(row=row+2, column=9).value
		sheet2.cell(row=line, column=10).value = sheet.cell(row=row+2, column=10).value
		sheet2.cell(row=line, column=11).value = sheet.cell(row=row+2, column=11).value
		sheet2.cell(row=line, column=12).value = sheet.cell(row=row+2, column=12).value
		sheet2.cell(row=line, column=13).value = sheet.cell(row=row+2, column=13).value
		sheet2.cell(row=line, column=14).value = sheet.cell(row=row+2, column=14).value
		sheet2.cell(row=line, column=15).value = sheet.cell(row=row+2, column=15).value
		sheet2.cell(row=line, column=16).value = sheet.cell(row=row+2, column=16).value
		sheet2.cell(row=line, column=17).value = sheet.cell(row=row+2, column=17).value


wb_final.save(filepath)
